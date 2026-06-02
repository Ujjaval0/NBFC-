from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


BASE_DIR = Path(r"C:\Users\freak\Videos\data")
OUTPUT_DIR = Path(r"C:\Users\freak\Desktop\random\outputs\nbfc_dashboard")
OUTPUT_PATH = OUTPUT_DIR / "NBFC_Dynamic_Dashboard.xlsx"


BRAND_NAVY = "1D1D1F"   # Apple Dark Slate
BRAND_BLUE = "0071E3"   # Apple Blue
BRAND_TEAL = "86868B"   # Apple Gray
BRAND_GREEN = "34C759"  # Apple Green
BRAND_GOLD = "FF9F0A"   # Apple Orange
BRAND_RED = "FF453A"    # Apple Red
LIGHT_BLUE = "FFFFFF"   # Clean White Cards
LIGHT_TEAL = "FFFFFF"
LIGHT_GREEN = "FFFFFF"
LIGHT_GOLD = "FFFFFF"
LIGHT_RED = "FFFFFF"
LIGHT_GRAY = "F5F5F7"   # Clean Apple Light Gray
WHITE = "FFFFFF"
TEXT = "1D1D1F"         # Dark contrast text
BORDER = "E8E8ED"       # Very soft border line


def ticket_bucket(amount: float) -> str:
    if amount < 300_000:
        return "Small"
    if amount < 750_000:
        return "Medium"
    if amount < 1_500_000:
        return "Large"
    return "High Value"


def cibil_band(score: int) -> str:
    if score < 650:
        return "<650"
    if score < 700:
        return "650-699"
    if score < 750:
        return "700-749"
    return "750+"


def income_band(income: int) -> str:
    if income < 50_000:
        return "<50k"
    if income < 100_000:
        return "50k-99k"
    if income < 150_000:
        return "100k-149k"
    return "150k+"


def customer_segment(row) -> str:
    if row["CIBIL_Score"] >= 750 and row["Monthly_Income"] >= 100_000:
        return "Prime"
    if row["CIBIL_Score"] >= 700 and row["Monthly_Income"] >= 50_000:
        return "Mass Affluent"
    if row["CIBIL_Score"] < 650:
        return "Risk Watch"
    return "Mass Market"


def collection_status(row) -> str:
    due = row.get("Amount_Due_Current_Month", 0)
    paid = row.get("Amount_Paid_Current_Month", 0)
    if due == 0:
        return "No Current Due"
    if paid >= due:
        return "Paid"
    if paid > 0:
        return "Partially Paid"
    return "Unpaid"


def load_data():
    apps = pd.read_csv(BASE_DIR / "Applications_and_Loans.csv")
    ops = pd.read_csv(BASE_DIR / "Daily_Operational_Logs.csv")
    repay = pd.read_csv(BASE_DIR / "Repayments_and_Performance.csv")

    apps["Application_Date"] = pd.to_datetime(apps["Application_Date"])
    ops["Stage_Start_Timestamp"] = pd.to_datetime(ops["Stage_Start_Timestamp"])
    ops["Stage_End_Timestamp"] = pd.to_datetime(ops["Stage_End_Timestamp"], errors="coerce")
    ops["Stage_Date"] = ops["Stage_Start_Timestamp"].dt.date
    ops["Stage_TAT_Hours"] = (
        ops["Stage_End_Timestamp"] - ops["Stage_Start_Timestamp"]
    ).dt.total_seconds() / 3600
    ops["Stage_TAT_Hours"] = ops["Stage_TAT_Hours"].round(2)

    apps["Ticket_Size_Bucket"] = apps["Requested_Amount"].apply(ticket_bucket)
    apps["Customer_Segment"] = apps.apply(customer_segment, axis=1)
    apps["CIBIL_Band"] = apps["CIBIL_Score"].apply(cibil_band)
    apps["Income_Band"] = apps["Monthly_Income"].apply(income_band)
    apps["Month"] = apps["Application_Date"].dt.to_period("M").dt.to_timestamp()

    assigned = (
        ops.groupby("Application_ID")["Assigned_To"]
        .apply(lambda s: ", ".join(sorted(set(s.dropna()))))
        .rename("Assigned_Users")
    )

    repay = repay.copy()
    repay["NPA_Flag"] = repay["NPA_Status"].str.contains("NPA", case=False).astype(int)
    repay["Collection_Status"] = repay.apply(collection_status, axis=1)

    model = apps.merge(assigned, on="Application_ID", how="left")
    model = model.merge(repay, on="Application_ID", how="left")
    model["Assigned_Users"] = model["Assigned_Users"].fillna("Unassigned")
    for col in [
        "Principal_Outstanding",
        "Amount_Due_Current_Month",
        "Amount_Paid_Current_Month",
        "DPD",
        "NPA_Flag",
        "Current_ROI",
        "Total_Tenure_Months",
    ]:
        model[col] = model[col].fillna(0)
    model["Loan_ID"] = model["Loan_ID"].fillna("")
    model["NPA_Status"] = model["NPA_Status"].fillna("No Active Loan")
    model["Collection_Status"] = model["Collection_Status"].fillna("No Active Loan")
    model["Has_Loan"] = (model["Loan_ID"] != "").astype(int)

    ops_model = ops.merge(
        apps[
            [
                "Application_ID",
                "Branch",
                "Loan_Type",
                "Ticket_Size_Bucket",
                "Customer_Segment",
                "Application_Status",
            ]
        ],
        on="Application_ID",
        how="left",
    )
    return apps, ops, repay, model, ops_model


def safe_excel_value(value):
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def add_dataframe_sheet(wb, name, df, table_name, freeze="A2"):
    ws = wb.create_sheet(name)
    headers = list(df.columns)
    ws.append(headers)
    for row in df.itertuples(index=False):
        ws.append([safe_excel_value(value) for value in row])

    end_col = get_column_letter(len(headers))
    end_row = len(df) + 1
    table = Table(displayName=table_name, ref=f"A1:{end_col}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BRAND_NAVY)
        cell.alignment = Alignment(horizontal="center")
    for idx, col in enumerate(headers, 1):
        width = min(max(len(str(col)) + 2, 12), 28)
        ws.column_dimensions[get_column_letter(idx)].width = width
    for col_name in ["Application_Date", "Month", "Stage_Start_Timestamp", "Stage_End_Timestamp", "Stage_Date"]:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            for row_idx in range(2, end_row + 1):
                ws.cell(row_idx, col_idx).number_format = "yyyy-mm-dd"
    return ws


def add_lists_sheet(wb, apps, ops, model):
    ws = wb.create_sheet("Lists")
    lists = {
        "Branch": ["All"] + sorted(apps["Branch"].dropna().unique().tolist()),
        "Loan_Type": ["All"] + sorted(apps["Loan_Type"].dropna().unique().tolist()),
        "Ticket_Size": ["All", "Small", "Medium", "Large", "High Value"],
        "Assigned_To": ["All"] + sorted(ops["Assigned_To"].dropna().unique().tolist()),
        "Customer_Segment": ["All"] + sorted(model["Customer_Segment"].dropna().unique().tolist()),
        "CIBIL_Band": ["<650", "650-699", "700-749", "750+"],
        "Stage": ["Lead_Generation", "Credit_Review", "Document_Verification", "Disbursal_Queue"],
        "NPA_Status": ["Standard", "Sub-Standard (NPA)", "Doubtful (NPA)"],
        "Months": sorted(model["Month"].dropna().unique().tolist()),
    }
    for col_idx, (label, values) in enumerate(lists.items(), 1):
        ws.cell(1, col_idx, label)
        ws.cell(1, col_idx).font = Font(bold=True, color=WHITE)
        ws.cell(1, col_idx).fill = PatternFill("solid", fgColor=BRAND_NAVY)
        for row_idx, value in enumerate(values, 2):
            ws.cell(row_idx, col_idx, safe_excel_value(value))
            if label == "Months":
                ws.cell(row_idx, col_idx).number_format = "mmm-yy"
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    ws.sheet_state = "hidden"
    return ws, lists


def model_conditions(extra=None):
    conditions = [
        "--(tblModel[Application_Date]>=Dashboard!$A$5)",
        "--(tblModel[Application_Date]<=Dashboard!$B$5)",
        'IF(Dashboard!$C$5="All",1,--(tblModel[Branch]=Dashboard!$C$5))',
        'IF(Dashboard!$D$5="All",1,--(tblModel[Loan_Type]=Dashboard!$D$5))',
        'IF(Dashboard!$E$5="All",1,--(tblModel[Ticket_Size_Bucket]=Dashboard!$E$5))',
        'IF(Dashboard!$F$5="All",1,--ISNUMBER(SEARCH(Dashboard!$F$5,tblModel[Assigned_Users])))',
        'IF(Dashboard!$G$5="All",1,--(tblModel[Customer_Segment]=Dashboard!$G$5))',
    ]
    if extra:
        conditions.extend(extra)
    return ",".join(conditions)


def ops_conditions(extra=None):
    conditions = [
        "--(tblOps[Stage_Date]>=Dashboard!$A$5)",
        "--(tblOps[Stage_Date]<=Dashboard!$B$5)",
        'IF(Dashboard!$C$5="All",1,--(tblOps[Branch]=Dashboard!$C$5))',
        'IF(Dashboard!$D$5="All",1,--(tblOps[Loan_Type]=Dashboard!$D$5))',
        'IF(Dashboard!$E$5="All",1,--(tblOps[Ticket_Size_Bucket]=Dashboard!$E$5))',
        'IF(Dashboard!$F$5="All",1,--(tblOps[Assigned_To]=Dashboard!$F$5))',
        'IF(Dashboard!$G$5="All",1,--(tblOps[Customer_Segment]=Dashboard!$G$5))',
    ]
    if extra:
        conditions.extend(extra)
    return ",".join(conditions)


def sumproduct(conditions, measure=None):
    parts = conditions
    if measure:
        parts = f"{parts},{measure}"
    return f"=SUMPRODUCT({parts})"


def avg_formula(conditions, measure, count_condition=None):
    denominator_conditions = conditions
    if count_condition:
        denominator_conditions = f"{denominator_conditions},{count_condition}"
    return f"=IFERROR(SUMPRODUCT({conditions},{measure})/SUMPRODUCT({denominator_conditions}),0)"


def style_title(ws, cell_range, title):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = title
    cell.font = Font(bold=True, color=WHITE, size=12)
    cell.fill = PatternFill("solid", fgColor=BRAND_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def add_calc_sheet(wb, model):
    ws = wb.create_sheet("Calc")
    ws.sheet_view.showGridLines = False

    total = model_conditions()
    approved = model_conditions(['--((tblModel[Application_Status]="Sanctioned")+(tblModel[Application_Status]="Disbursed")>0)'])
    disbursed = model_conditions(['--(tblModel[Application_Status]="Disbursed")'])
    rejected = model_conditions(['--(tblModel[Application_Status]="Rejected")'])
    loans = model_conditions(['--(tblModel[Has_Loan]=1)'])
    npa = model_conditions(['--(tblModel[NPA_Flag]=1)'])

    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    funnel_rows = [
        ("Applications", sumproduct(total)),
        ("Sanctioned/Approved", sumproduct(approved)),
        ("Disbursed", sumproduct(disbursed)),
    ]
    for idx, (label, formula) in enumerate(funnel_rows, 2):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, formula)

    ws["D1"] = "Month"
    ws["E1"] = "Applications"
    ws["F1"] = "Disbursed Loans"
    ws["G1"] = "Sanctioned Amount"
    months = sorted(model["Month"].dropna().unique())
    for row_idx, month in enumerate(months, 2):
        ws.cell(row_idx, 4, safe_excel_value(month))
        ws.cell(row_idx, 4).number_format = "mmm-yy"
        month_start = f"--(tblModel[Month]=$D{row_idx})"
        ws.cell(row_idx, 5, sumproduct(model_conditions([month_start])))
        ws.cell(row_idx, 6, sumproduct(model_conditions([month_start, '--(tblModel[Application_Status]="Disbursed")'])))
        ws.cell(row_idx, 7, sumproduct(model_conditions([month_start]), "tblModel[Sanctioned_Amount]"))
        ws.cell(row_idx, 7).number_format = '₹#,##0'

    ws["I1"] = "NPA Status"
    ws["J1"] = "Loans"
    for row_idx, status in enumerate(["Standard", "Sub-Standard (NPA)", "Doubtful (NPA)"], 2):
        ws.cell(row_idx, 9, status)
        ws.cell(row_idx, 10, sumproduct(model_conditions([f'--(tblModel[NPA_Status]="{status}")'])))

    ws["L1"] = "Branch"
    ws["M1"] = "Applications"
    ws["N1"] = "Approval Rate"
    for row_idx, branch in enumerate(sorted(model["Branch"].dropna().unique()), 2):
        ws.cell(row_idx, 12, branch)
        branch_total = model_conditions([f"--(tblModel[Branch]=$L{row_idx})"])
        branch_approved = model_conditions([
            f"--(tblModel[Branch]=$L{row_idx})",
            '--((tblModel[Application_Status]="Sanctioned")+(tblModel[Application_Status]="Disbursed")>0)',
        ])
        ws.cell(row_idx, 13, sumproduct(branch_total))
        ws.cell(row_idx, 14, f"=IFERROR(SUMPRODUCT({branch_approved})/SUMPRODUCT({branch_total}),0)")
        ws.cell(row_idx, 14).number_format = "0.0%"

    ws["P1"] = "Loan Type"
    ws["Q1"] = "Applications"
    for row_idx, loan_type in enumerate(sorted(model["Loan_Type"].dropna().unique()), 2):
        ws.cell(row_idx, 16, loan_type)
        ws.cell(row_idx, 17, sumproduct(model_conditions([f"--(tblModel[Loan_Type]=$P{row_idx})"])))

    ws["S1"] = "CIBIL Band"
    ws["T1"] = "Approval Rate"
    for row_idx, band in enumerate(["<650", "650-699", "700-749", "750+"], 2):
        ws.cell(row_idx, 19, band)
        band_total = model_conditions([f"--(tblModel[CIBIL_Band]=$S{row_idx})"])
        band_approved = model_conditions([
            f"--(tblModel[CIBIL_Band]=$S{row_idx})",
            '--((tblModel[Application_Status]="Sanctioned")+(tblModel[Application_Status]="Disbursed")>0)',
        ])
        ws.cell(row_idx, 20, f"=IFERROR(SUMPRODUCT({band_approved})/SUMPRODUCT({band_total}),0)")
        ws.cell(row_idx, 20).number_format = "0.0%"

    ws["V1"] = "Stage"
    ws["W1"] = "Avg TAT Hours"
    ws["X1"] = "Pending Cases"
    for row_idx, stage in enumerate(["Lead_Generation", "Credit_Review", "Document_Verification", "Disbursal_Queue"], 2):
        ws.cell(row_idx, 22, stage)
        stage_conditions = ops_conditions([f"--(tblOps[Stage]=$V{row_idx})", '--(tblOps[Stage_TAT_Hours]>0)'])
        ws.cell(row_idx, 23, avg_formula(stage_conditions, "tblOps[Stage_TAT_Hours]"))
        ws.cell(row_idx, 23).number_format = "0.0"
        ws.cell(row_idx, 24, sumproduct(ops_conditions([f"--(tblOps[Stage]=$V{row_idx})", '--(tblOps[Status]="Pending")'])))

    ws["Z1"] = "Customer Segment"
    ws["AA1"] = "Applications"
    ws["AB1"] = "NPA Rate"
    for row_idx, segment in enumerate(sorted(model["Customer_Segment"].dropna().unique()), 2):
        ws.cell(row_idx, 26, segment)
        segment_total = model_conditions([f"--(tblModel[Customer_Segment]=$Z{row_idx})"])
        segment_npa = model_conditions([
            f"--(tblModel[Customer_Segment]=$Z{row_idx})",
            "--(tblModel[NPA_Flag]=1)",
        ])
        segment_loans = model_conditions([
            f"--(tblModel[Customer_Segment]=$Z{row_idx})",
            "--(tblModel[Has_Loan]=1)",
        ])
        ws.cell(row_idx, 27, sumproduct(segment_total))
        ws.cell(row_idx, 28, f"=IFERROR(SUMPRODUCT({segment_npa})/SUMPRODUCT({segment_loans}),0)")
        ws.cell(row_idx, 28).number_format = "0.0%"

    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value:
                cell.font = Font(bold=True, color=WHITE)
                cell.fill = PatternFill("solid", fgColor=BRAND_NAVY)
                cell.alignment = Alignment(horizontal="center")
    for col in range(1, 29):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.sheet_state = "hidden"
    return ws


def add_card(ws, cell_range, title, formula, number_format, fill):
    start, end = cell_range.split(":")
    start_cell = ws[start]
    end_cell = ws[end]
    start_row, end_row = start_cell.row, end_cell.row
    start_col, end_col = start_cell.column, end_cell.column
    title_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{start_row}"
    value_range = f"{get_column_letter(start_col)}{start_row + 1}:{get_column_letter(end_col)}{end_row}"

    ws.merge_cells(title_range)
    ws.merge_cells(value_range)

    title_cell = ws.cell(start_row, start_col)
    title_cell.value = title
    title_cell.font = Font(bold=True, color=BRAND_TEAL, size=9)
    title_cell.fill = PatternFill("solid", fgColor=fill)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    value_cell = ws.cell(start_row + 1, start_col)
    value_cell.value = formula
    value_cell.number_format = number_format
    value_cell.font = Font(bold=True, color=BRAND_NAVY, size=16)
    value_cell.fill = PatternFill("solid", fgColor=fill)
    value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for c in row:
            c.fill = PatternFill("solid", fgColor=fill)
            c.border = Border(
                left=Side(style="thin", color=BORDER),
                right=Side(style="thin", color=BORDER),
                top=Side(style="thin", color=BORDER),
                bottom=Side(style="thin", color=BORDER),
            )


def add_dashboard(wb, apps, model, lists):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for row in range(1, 91):
        ws.row_dimensions[row].height = 22

    ws.merge_cells("A1:N1")
    ws["A1"] = "NBFC Decision Dashboard"
    ws["A1"].font = Font(bold=True, color=WHITE, size=20)
    ws["A1"].fill = PatternFill("solid", fgColor=BRAND_NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:N2")
    ws["A2"] = "Dynamic dashboard for Sales, Credit, Operations, and Management. Change the filter values below to recalculate KPIs and charts."
    ws["A2"].font = Font(color=TEXT, italic=True)
    ws["A2"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    style_title(ws, "A3:G3", "Dashboard Filters")
    filters = [
        ("Start Date", apps["Application_Date"].min()),
        ("End Date", apps["Application_Date"].max()),
        ("Branch", "All"),
        ("Loan Type", "All"),
        ("Ticket Size", "All"),
        ("User / Assigned To", "All"),
        ("Customer Segment", "All"),
    ]
    for idx, (label, value) in enumerate(filters, 1):
        # Label in Row 4
        cell_lbl = ws.cell(4, idx, label)
        cell_lbl.font = Font(bold=True, color=TEXT)
        cell_lbl.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        cell_lbl.alignment = Alignment(horizontal="center", vertical="center")

        # Value in Row 5
        cell_val = ws.cell(5, idx, safe_excel_value(value))
        cell_val.fill = PatternFill("solid", fgColor=WHITE)
        cell_val.border = Border(
            left=Side(style="thin", color=BORDER),
            right=Side(style="thin", color=BORDER),
            top=Side(style="thin", color=BORDER),
            bottom=Side(style="thin", color=BRAND_BLUE)
        )
        cell_val.alignment = Alignment(horizontal="center", vertical="center")
        if "Date" in label:
            cell_val.number_format = "yyyy-mm-dd"

    validation_ranges = {
        "C5": ("Lists", 1, len(lists["Branch"]) + 1),
        "D5": ("Lists", 2, len(lists["Loan_Type"]) + 1),
        "E5": ("Lists", 3, len(lists["Ticket_Size"]) + 1),
        "F5": ("Lists", 4, len(lists["Assigned_To"]) + 1),
        "G5": ("Lists", 5, len(lists["Customer_Segment"]) + 1),
    }
    for cell, (sheet, col, max_row) in validation_ranges.items():
        col_letter = get_column_letter(col)
        dv = DataValidation(
            type="list",
            formula1=f"='{sheet}'!${col_letter}$2:${col_letter}${max_row}",
            allow_blank=False,
        )
        ws.add_data_validation(dv)
        dv.add(ws[cell])

    ws.merge_cells("H3:N5")
    ws["H3"] = "Sales, Credit, Operations & Portfolio Management Overview"
    ws["H3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["H3"].font = Font(color=BRAND_NAVY, italic=True, size=11, bold=True)
    ws["H3"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    total = model_conditions()
    approved = model_conditions(['--((tblModel[Application_Status]="Sanctioned")+(tblModel[Application_Status]="Disbursed")>0)'])
    disbursed = model_conditions(['--(tblModel[Application_Status]="Disbursed")'])
    rejected = model_conditions(['--(tblModel[Application_Status]="Rejected")'])
    loans = model_conditions(['--(tblModel[Has_Loan]=1)'])
    npa = model_conditions(['--(tblModel[NPA_Flag]=1)'])

    cards = [
        ("A7:C9", "Total Applications", sumproduct(total), "#,##0", LIGHT_BLUE),
        ("D7:F9", "Disbursed Loans", sumproduct(disbursed), "#,##0", LIGHT_GREEN),
        ("G7:I9", "Sanctioned Amount", sumproduct(model_conditions(), "tblModel[Sanctioned_Amount]"), '₹#,##0', LIGHT_GOLD),
        ("J7:L9", "Approval Rate", f"=IFERROR(SUMPRODUCT({approved})/SUMPRODUCT({total}),0)", "0.0%", LIGHT_GREEN),
        ("M7:N9", "Avg CIBIL", avg_formula(total, "tblModel[CIBIL_Score]"), "0", LIGHT_BLUE),
        ("A11:C13", "Rejection Rate", f"=IFERROR(SUMPRODUCT({rejected})/SUMPRODUCT({total}),0)", "0.0%", LIGHT_RED),
        ("D11:F13", "Avg Ticket Size", avg_formula(model_conditions(['--(tblModel[Sanctioned_Amount]>0)']), "tblModel[Sanctioned_Amount]"), '₹#,##0', LIGHT_GOLD),
        ("G11:I13", "Principal Outstanding", sumproduct(model_conditions(), "tblModel[Principal_Outstanding]"), '₹#,##0', LIGHT_BLUE),
        ("J11:L13", "Collection Efficiency", f"=IFERROR(SUMPRODUCT({model_conditions()},tblModel[Amount_Paid_Current_Month])/SUMPRODUCT({model_conditions()},tblModel[Amount_Due_Current_Month]),0)", "0.0%", LIGHT_TEAL),
        ("M11:N13", "NPA %", f"=IFERROR(SUMPRODUCT({npa})/SUMPRODUCT({loans}),0)", "0.0%", LIGHT_RED),
    ]
    for card in cards:
        add_card(ws, *card)

    chart_specs = [
        ("A15:E31", "Application Funnel", "Funnel: Applications to disbursal"),
        ("F15:J31", "Monthly Trend", "Application and disbursement count by month"),
        ("K15:N31", "Segment Risk", "NPA % by Customer Segment"),
        ("A33:E49", "Branch Performance", "Applications by branch"),
        ("F33:J49", "Loan Type Mix", "Applications by loan type"),
        ("K33:N49", "Credit Quality", "Approval rate by CIBIL band"),
        ("A51:E67", "Operations TAT", "Average stage turnaround time"),
        ("A69:E85", "Customer Segments", "Applications by customer segment"),
    ]
    for cell_range, title, subtitle in chart_specs:
        start, end = cell_range.split(":")
        start_cell = ws[start]
        end_cell = ws[end]
        title_range = f"{get_column_letter(start_cell.column)}{start_cell.row}:{get_column_letter(end_cell.column)}{start_cell.row}"
        style_title(ws, title_range, title)
        top = start_cell.row + 1
        left = start_cell.column
        ws.cell(top, left, subtitle)
        ws.cell(top, left).font = Font(italic=True, color="666666")

    style_title(ws, "F51:N51", "Pending Operations")
    ws["F52"] = "Application"
    ws["G52"] = "Stage"
    ws["H52"] = "Assigned To"
    ws["I52"] = "Stage Start"
    ws["J52"] = "Status"
    for col in range(6, 11):
        cell = ws.cell(52, col)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    ws["F53"] = (
        '=FILTER(CHOOSECOLS(tblOps,2,3,4,5,7),'
        '(tblOps[Status]="Pending")*'
        '(tblOps[Stage_Date]>=Dashboard!$A$5)*'
        '(tblOps[Stage_Date]<=Dashboard!$B$5)*'
        'IF(Dashboard!$C$5="All",1,--(tblOps[Branch]=Dashboard!$C$5))*'
        'IF(Dashboard!$D$5="All",1,--(tblOps[Loan_Type]=Dashboard!$D$5))*'
        'IF(Dashboard!$E$5="All",1,--(tblOps[Ticket_Size_Bucket]=Dashboard!$E$5))*'
        'IF(Dashboard!$F$5="All",1,--(tblOps[Assigned_To]=Dashboard!$F$5))*'
        'IF(Dashboard!$G$5="All",1,--(tblOps[Customer_Segment]=Dashboard!$G$5)),'
        '"No pending cases")'
    )

    ws.freeze_panes = "A7"
    return ws


def set_series_color(chart, index, color_hex):
    try:
        if len(chart.series) > index:
            series = chart.series[index]
            if not hasattr(series, "graphicalProperties") or series.graphicalProperties is None:
                from openpyxl.chart.shapes import GraphicalProperties
                series.graphicalProperties = GraphicalProperties()
            series.graphicalProperties.solidFill = color_hex
    except Exception:
        pass


def set_line_color(chart, index, color_hex):
    try:
        if len(chart.series) > index:
            series = chart.series[index]
            if not hasattr(series, "graphicalProperties") or series.graphicalProperties is None:
                from openpyxl.chart.shapes import GraphicalProperties
                series.graphicalProperties = GraphicalProperties()
            from openpyxl.drawing.line import LineProperties
            series.graphicalProperties.line = LineProperties()
            series.graphicalProperties.line.solidFill = color_hex
            series.graphicalProperties.line.width = 25000  # 2.5pt
    except Exception:
        pass


def add_charts(wb):
    dashboard = wb["Dashboard"]
    calc = wb["Calc"]

    # 1. Application Funnel
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Application Funnel"
    chart.y_axis.title = "Stage"
    chart.x_axis.title = "Count"
    chart.add_data(Reference(calc, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    chart.set_categories(Reference(calc, min_col=1, min_row=2, max_row=4))
    chart.height = 7.5
    chart.width = 9
    chart.legend = None
    set_series_color(chart, 0, BRAND_BLUE)
    dashboard.add_chart(chart, "A18")

    # 2. Monthly Trend
    chart = LineChart()
    chart.title = "Monthly Applications"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Month"
    chart.add_data(Reference(calc, min_col=5, max_col=6, min_row=1, max_row=11), titles_from_data=True)
    chart.set_categories(Reference(calc, min_col=4, min_row=2, max_row=11))
    chart.height = 7.5
    chart.width = 9
    set_line_color(chart, 0, BRAND_NAVY)
    set_line_color(chart, 1, BRAND_TEAL)
    dashboard.add_chart(chart, "F18")

    # 3. Segment Risk (replaces Portfolio Risk Mix Doughnut)
    chart = BarChart()
    chart.type = "col"
    chart.title = "NPA % by Segment"
    chart.y_axis.title = "NPA %"
    chart.add_data(Reference(calc, min_col=28, min_row=1, max_row=5), titles_from_data=True)
    chart.set_categories(Reference(calc, min_col=26, min_row=2, max_row=5))
    chart.height = 7.5
    chart.width = 7
    chart.legend = None
    set_series_color(chart, 0, BRAND_RED)
    dashboard.add_chart(chart, "K18")

    # 4. Branch-wise Applications
    chart = BarChart()
    chart.type = "col"
    chart.title = "Branch-wise Applications"
    chart.y_axis.title = "Applications"
    chart.add_data(Reference(calc, min_col=13, min_row=1, max_row=5), titles_from_data=True)
    chart.set_categories(Reference(calc, min_col=12, min_row=2, max_row=5))
    chart.height = 7.5
    chart.width = 9
    chart.legend = None
    set_series_color(chart, 0, BRAND_NAVY)
    dashboard.add_chart(chart, "A36")

    # 5. Loan Type Mix
    chart = BarChart()
    chart.type = "col"
    chart.title = "Loan Type Mix"
    chart.y_axis.title = "Applications"
    chart.add_data(Reference(calc, min_col=17, min_row=1, max_row=5), titles_from_data=True)
    chart.set_categories(Reference(calc, min_col=16, min_row=2, max_row=5))
    chart.height = 7.5
    chart.width = 9
    chart.legend = None
    set_series_color(chart, 0, BRAND_BLUE)
    dashboard.add_chart(chart, "F36")

    # 6. Credit Quality
    chart = BarChart()
    chart.type = "col"
    chart.title = "Approval Rate by CIBIL Band"
    chart.y_axis.title = "Approval Rate"
    chart.add_data(Reference(calc, min_col=20, min_row=1, max_row=5), titles_from_data=True)
    chart.set_categories(Reference(calc, min_col=19, min_row=2, max_row=5))
    chart.height = 7.5
    chart.width = 7
    chart.legend = None
    set_series_color(chart, 0, BRAND_TEAL)
    dashboard.add_chart(chart, "K36")

    # 7. Average Stage TAT
    chart = BarChart()
    chart.type = "col"
    chart.title = "Average Stage TAT"
    chart.y_axis.title = "Hours"
    chart.add_data(Reference(calc, min_col=23, min_row=1, max_row=5), titles_from_data=True)
    chart.set_categories(Reference(calc, min_col=22, min_row=2, max_row=5))
    chart.height = 7.5
    chart.width = 9
    set_series_color(chart, 0, BRAND_TEAL)
    set_series_color(chart, 1, BRAND_GOLD)
    dashboard.add_chart(chart, "A54")

    # 8. Segment-wise Applications
    chart = BarChart()
    chart.type = "col"
    chart.title = "Segment-wise Applications"
    chart.y_axis.title = "Applications"
    chart.add_data(Reference(calc, min_col=27, min_row=1, max_row=5), titles_from_data=True)
    chart.set_categories(Reference(calc, min_col=26, min_row=2, max_row=5))
    chart.height = 7.5
    chart.width = 9
    chart.legend = None
    set_series_color(chart, 0, BRAND_NAVY)
    dashboard.add_chart(chart, "A72")


def add_kpi_doc(wb):
    ws = wb.create_sheet("KPI Documentation")
    ws.sheet_view.showGridLines = False
    headers = ["KPI", "Persona", "Reason for Selection", "Business Decision Supported", "Visualization Logic"]
    ws.append(headers)
    rows = [
        ("Total Applications", "Management, Sales", "Shows demand entering the lending funnel.", "Plan staffing, branch targets, and acquisition campaigns.", "KPI card and monthly line chart."),
        ("Disbursed Loans", "Management, Sales", "Measures completed conversion to active loans.", "Assess revenue pipeline and sales execution quality.", "KPI card and funnel stage."),
        ("Requested Amount", "Management", "Captures gross demand value.", "Compare demand value against sanctioned/disbursed value.", "Calculated in model and available for analysis."),
        ("Sanctioned Amount", "Management, Credit", "Shows approved lending value.", "Track credit-approved business volume.", "KPI card and monthly summary."),
        ("Approval Rate", "Management, Credit", "Balances growth with underwriting selectivity.", "Tune credit policy and branch coaching.", "KPI card and CIBIL-band bar chart."),
        ("Rejection Rate", "Credit, Sales", "Highlights application quality and risk filtering.", "Identify branches or loan products needing better sourcing.", "KPI card and filtered comparisons."),
        ("Average Ticket Size", "Management, Sales", "Indicates value per approved loan.", "Set product strategy and segment focus.", "KPI card."),
        ("Principal Outstanding", "Management", "Represents active portfolio exposure.", "Monitor portfolio scale and capital exposure.", "KPI card."),
        ("Collection Efficiency", "Management, Operations", "Measures current-month repayment realization.", "Escalate collections focus or operations follow-up.", "KPI card."),
        ("NPA %", "Management, Credit", "Tracks risky/delinquent portfolio share.", "Adjust underwriting, collection intensity, and risk appetite.", "KPI card and risk mix chart."),
        ("Application Funnel", "Sales, Management", "Shows drop-offs across application, approval, and disbursal.", "Prioritize funnel stage improvements.", "Horizontal bar funnel."),
        ("Branch-wise Applications", "Sales, Management", "Compares demand by geography.", "Allocate targets, incentives, and staffing.", "Bar chart."),
        ("Loan Type Mix", "Sales, Management", "Shows product demand split.", "Refine product/channel strategy.", "Bar chart."),
        ("Segment-wise Applications", "Sales, Credit", "Connects demand to customer quality/segment.", "Balance growth segments with risk appetite.", "Bar chart."),
        ("Average CIBIL Score", "Credit", "Summarizes applicant credit quality.", "Monitor sourcing quality and risk profile.", "KPI card."),
        ("Approval Rate by CIBIL Band", "Credit", "Tests whether approvals align with risk bands.", "Audit underwriting and policy thresholds.", "Bar chart."),
        ("Rejection Hotspots", "Credit, Sales", "No rejection reason exists, so branch/product/score patterns are used.", "Find where sourcing quality or policy friction is highest.", "Filterable rejection-rate KPI."),
        ("Average Stage TAT", "Operations", "Shows processing speed by workflow stage.", "Identify bottlenecks and SLA risk.", "Stage bar chart."),
        ("Pending Operational Cases", "Operations", "Lists work still stuck in process.", "Drive daily follow-up and ownership.", "Dynamic filtered table."),
        ("Bottleneck Stage", "Operations, Management", "Highest TAT stage indicates process constraint.", "Focus process improvement and staffing.", "Stage TAT chart."),
    ]
    for row in rows:
        ws.append(row)
    table = Table(displayName="tblKPIDoc", ref=f"A1:E{len(rows)+1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)
    widths = [26, 24, 44, 48, 34]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=1, max_row=len(rows) + 1, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BRAND_NAVY)
    ws.freeze_panes = "A2"


def apply_workbook_styles(wb):
    for ws in wb.worksheets:
        if ws.title in {"Lists", "Calc"}:
            continue
        for row in ws.iter_rows():
            for cell in row:
                font = copy(cell.font)
                alignment = copy(cell.alignment)
                font.name = "Inter"
                if alignment.vertical is None:
                    alignment.vertical = "center"
                cell.font = font
                cell.alignment = alignment

    dashboard = wb["Dashboard"]
    for col in range(1, 8):
        # Header formatting in Row 4
        dashboard.cell(4, col).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        dashboard.cell(4, col).border = Border(
            left=Side(style="thin", color=BORDER),
            right=Side(style="thin", color=BORDER),
            top=Side(style="thin", color=BORDER),
            bottom=Side(style="thin", color=BORDER),
        )
        # Value cell formatting in Row 5
        dashboard.cell(5, col).fill = PatternFill("solid", fgColor=WHITE)
        dashboard.cell(5, col).border = Border(
            left=Side(style="thin", color=BORDER),
            right=Side(style="thin", color=BORDER),
            top=Side(style="thin", color=BORDER),
            bottom=Side(style="thin", color=BRAND_BLUE),
        )
    dashboard.conditional_formatting.add(
        "M11:N13",
        CellIsRule(operator="greaterThan", formula=["0.05"], fill=PatternFill("solid", fgColor="FFECEB")),
    )


def set_calc_mode(wb):
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"


def validate_outputs(apps, ops, repay, model, wb):
    assert len(apps) == 150, f"Expected 150 applications, found {len(apps)}"
    assert len(ops) == 150, f"Expected 150 operation logs, found {len(ops)}"
    assert len(repay) == 107, f"Expected 107 repayment rows, found {len(repay)}"
    assert model["Application_ID"].is_unique, "Application model should be one row per application"
    assert set(repay["Application_ID"]).issubset(set(apps["Application_ID"])), "Repayment IDs must match applications"
    for required in ["Dashboard", "KPI Documentation", "Data Model", "Applications", "Operations", "Repayments"]:
        assert required in wb.sheetnames, f"Missing sheet: {required}"
    assert len(wb["Dashboard"]._charts) == 8, "Expected 8 dashboard charts"


def main():
    apps, ops, repay, model, ops_model = load_data()

    wb = Workbook()
    wb.remove(wb.active)

    add_dataframe_sheet(wb, "Applications", apps, "tblApplications")
    add_dataframe_sheet(wb, "Operations", ops, "tblOperations")
    add_dataframe_sheet(wb, "Repayments", repay, "tblRepayments")
    add_dataframe_sheet(wb, "Data Model", model, "tblModel")
    add_dataframe_sheet(wb, "Operations Model", ops_model, "tblOps")
    wb["Operations Model"].sheet_state = "hidden"
    lists_ws, lists = add_lists_sheet(wb, apps, ops, model)
    add_calc_sheet(wb, model)
    add_dashboard(wb, apps, model, lists)
    add_charts(wb)
    add_kpi_doc(wb)
    apply_workbook_styles(wb)
    set_calc_mode(wb)

    validate_outputs(apps, ops, repay, model, wb)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)

    loaded = load_workbook(OUTPUT_PATH, data_only=False)
    validate_outputs(apps, ops, repay, model, loaded)
    formula_cells = [
        loaded["Dashboard"]["A8"].value,
        loaded["Dashboard"]["J8"].value,
        loaded["Calc"]["B2"].value,
        loaded["Dashboard"]["F53"].value,
    ]
    print("saved=", OUTPUT_PATH)
    print("sheets=", ", ".join(loaded.sheetnames))
    print("charts=", len(loaded["Dashboard"]._charts))
    print("formulas_present=", all(isinstance(v, str) and v.startswith("=") for v in formula_cells if v))
    print("rows=", {"applications": len(apps), "operations": len(ops), "repayments": len(repay), "model": len(model)})


if __name__ == "__main__":
    main()
